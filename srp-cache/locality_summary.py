"""
Scrapes the SRP Locality Summary report for each NC-215 Triangle locality
and writes results to a per-locality-tab Google Sheet.

Usage:
    python locality_summary.py

Set SRP_USERNAME, SRP_PASSWORD, and TOTP_SECRET in .env to authenticate
automatically. If any are missing the browser opens for manual login.
"""

import asyncio
import datetime
import io
from pathlib import Path

import openpyxl
from playwright.async_api import async_playwright

from srp_lib.auth import load_credentials, authenticate, totp_watchdog
from srp_lib.browser import launch_browser, download_excel
from srp_lib.sheets import open_sheet

SHEET_ID    = '1wV1kZZd-vkxfIEqU-PlEl5aMPr6Bk2c7VOvKiyxxv5k'
EXPORTS_DIR          = Path(__file__).parent / 'exports'
TIMEOUT              = 60_000
HEADER_ROWS          = 3       # rows 1–3 are headers; data starts at row 4
DATE_COL             = 4       # column E, 0-indexed
INTER_LOCALITY_DELAY = 10_000  # ms between locality requests; reduces server load
MAX_RETRIES          = 2

LOCALITIES = [
    'Apex', 'Carrboro', 'Cary', 'Chapel Hill',
    'Durham', 'Durham County', 'Fuquay-Varina', 'Garner',
    'Hillsborough', 'Holly Springs', 'Knightdale',
    'Morrisville', 'Orange County', 'Raleigh', 'Rolesville',
    'Wake County', 'Wake Forest', 'Wendell', 'Zebulon',
]

# Combined 72-column header: first 5 from Row 1, rest from Row 3 leaf headers.
EXPECTED_COLUMNS = [
    'Locality', 'Cluster', 'Region', 'National Community', 'Date as of',
    'Book 1', 'Book 2', 'Book 3 (G1)', 'Book 3 (G2)', 'Book 3 (G3)',
    'Book 3 (G4)', 'Book 3 (G5)', 'Book 4', 'Book 5', 'Book 5 BR1',
    'Book 5 BR2', 'Book 5 BR3', 'Book 6', 'Book 7', 'Book 7 BR1',
    'Book 7 BR2', 'Book 8 (U1)', 'Book 8 (U2)', 'Book 8 (U3)',
    'Book 9 (U1)', 'Book 9 (U2)', 'Book 9 (U3)', 'Book 10 (U1)',
    'Book 10 (U2)', 'Book 10 (U3)', 'Book 11 (U1)', 'Book 11 (U2)',
    'Book 11 (U3)', 'Book 12 (U1)', 'Book 12 (U2)', 'Book 12 (U3)',
    'Book 13 (U1)', 'Book 13 (U2)', 'Book 13 (U3)', 'Book 14 (U1)',
    'Book 14 (U2)', 'Book 14 (U3)',
    'No.', 'Att.', 'Friends of the Faith',
    'No.', 'Att.', 'Friends of the Faith',
    'No.', 'Att.', 'Friends of the Faith',
    'No.', 'Att.', 'Friends of the Faith',
    'No.', 'Att.', 'Friends of the Faith',
    'No.', 'Att.', 'Friends of the Faith',
    'Children', 'Junior Youth', 'Youth', 'Adult Men', 'Adult Women',
    'Total Believers', 'Has Home Visits', 'No. of Homes Visited for Deepening',
    'Has 19 Day Feast', 'Att. at 19 Day Feast', 'Has Holy Days',
    'Att. at Holy Days',
]


def validate_columns(actual: list, expected: list) -> str | None:
    """Return an error string describing any schema mismatch, or None if columns match.

    Uses positional comparison so repeated column names (e.g. 'No.', 'Att.')
    are distinguished by their index rather than treated as identical.
    """
    if actual == expected:
        return None
    lines = ['Column mismatch — aborting.']
    if len(actual) != len(expected):
        lines.append(f'  Expected {len(expected)} columns, got {len(actual)}.')
    else:
        diffs = [
            (i, expected[i], actual[i])
            for i in range(len(expected))
            if actual[i] != expected[i]
        ]
        for idx, exp, got in diffs:
            lines.append(f'  Col {idx}: expected {exp!r}, got {got!r}')
    return '\n'.join(lines)


def extract_headers(xlsx_bytes: bytes) -> list:
    """Return the combined header list (Row 1 for cols 0–4, Row 3 for cols 5 onward)."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    row1 = list(rows[0])
    row3 = list(rows[2])
    return row1[:5] + row3[5:]


def parse_row(xlsx_bytes: bytes) -> list:
    """Return the single data row from a Locality Summary Excel as a list (None → '')."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    def _coerce(v):
        if v is None:
            return ''
        if isinstance(v, (datetime.datetime, datetime.date)):
            return v.strftime('%Y-%m-%d')
        return v

    return [_coerce(v) for v in rows[3]]


def parse_date(v) -> datetime.date:
    """Parse a date value into a datetime.date.

    Accepts:
    - datetime.datetime (returned by openpyxl for date cells)
    - datetime.date
    - str in ISO format '%Y-%m-%d' (produced by parse_row after coercion)
    - str in '%m/%d/%y' format (e.g. '5/3/26') from sheet reads
    """
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v)
    for fmt in ('%Y-%m-%d', '%m/%d/%y'):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    raise ValueError(f'Unrecognised date format: {s!r}')


def should_append(last_date_str: str | None, new_date_str: str) -> bool:
    """Return True when new_date_str represents a new SRP reporting period."""
    if last_date_str is None:
        return True
    return parse_date(new_date_str) != parse_date(last_date_str)


def write_locality_data(tab, row_values: list, new_date_val) -> None:
    """Overwrite the last data row if date matches, otherwise append a new row.

    Inlines last-date extraction to avoid a double get_all_values() round-trip.
    new_date_val may be a datetime.datetime, datetime.date, or str.
    """
    all_values = tab.get_all_values()
    data_rows = all_values[HEADER_ROWS:]

    last_date = None
    if data_rows:
        last_row = data_rows[-1]
        if len(last_row) > DATE_COL and last_row[DATE_COL]:
            last_date = last_row[DATE_COL]

    if should_append(last_date, new_date_val):
        next_row = HEADER_ROWS + len(data_rows) + 1  # 1-indexed sheet row
        tab.update(values=[row_values], range_name=f'A{next_row}')
        print(f'  Appended (Date as of: {new_date_val}).')
    else:
        last_row_num = HEADER_ROWS + len(data_rows)  # 1-indexed sheet row
        tab.update(values=[row_values], range_name=f'A{last_row_num}')
        print(f'  Overwrote last row (Date as of: {new_date_val}).')


async def navigate_to_locality_summary(page):
    """Navigate to Reports → Locality Summary from any page state (including home)."""
    await page.get_by_role('link', name='Reports').click()
    await page.wait_for_timeout(1_000)
    await page.get_by_role('button', name='Locality Summary').click()
    await page.wait_for_timeout(1_000)


async def scrape_locality(page, locality: str) -> bytes | None:
    """Change scope to locality, click Locality Summary, export Excel. Returns None on failure."""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            if attempt > 1:
                print(f'  Retrying {locality!r} (attempt {attempt})...')
                await navigate_to_locality_summary(page)

            # Open scope dropdown and select the locality (exact match avoids substring collisions)
            await page.locator('#dropdownLocationSelector').click()
            await page.wait_for_timeout(500)
            await page.locator(f'span.k-in:text-is("{locality}")').click()

            # Confirm scope changed before interacting further
            await page.locator('#dropdownLocationSelector').filter(has_text=locality).wait_for(
                state='visible', timeout=TIMEOUT
            )

            # SRP shows a loading modal while fetching locality data; some localities are slow
            await page.locator('.modal.app-modal-dlg-container.in').wait_for(
                state='hidden', timeout=300_000
            )

            # Locality Summary button only appears at locality scope
            await page.get_by_role('button', name='Locality Summary').click()
            await page.get_by_role('button', name='Export Data |').wait_for(
                state='visible', timeout=300_000
            )
            return await download_excel(page, EXPORTS_DIR, timeout=300_000)

        except Exception as e:
            short = str(e).split('\n')[0]
            if attempt < MAX_RETRIES:
                print(f'  WARNING: attempt {attempt} failed for {locality!r} ({short}) — re-navigating and retrying')
                try:
                    await navigate_to_locality_summary(page)
                except Exception:
                    pass
            else:
                print(f'  WARNING: skipping {locality!r} after {MAX_RETRIES} attempts ({short})')

    return None


async def main():
    creds = load_credentials()
    EXPORTS_DIR.mkdir(exist_ok=True)

    async with async_playwright() as pw:
        browser, context = await launch_browser(pw)
        page = await context.new_page()
        page.set_default_timeout(TIMEOUT)
        page.set_default_navigation_timeout(TIMEOUT)

        await page.goto(creds['url'])
        await page.wait_for_load_state('domcontentloaded')
        await page.wait_for_timeout(2_000)

        if 'login' in page.url.lower():
            if creds['username'] and creds['password'] and creds['totp_secret']:
                await authenticate(page, creds['username'], creds['password'], creds['totp_secret'])
            else:
                print('Not logged in. Set credentials in .env or log in manually and press Enter.')
                input()

        stop_event = asyncio.Event()
        watchdog = asyncio.create_task(totp_watchdog(page, stop_event, creds['totp_secret']))

        await page.get_by_role('link', name='Reports').click()
        await page.wait_for_timeout(500)

        locality_bytes: dict[str, bytes] = {}
        for i, locality in enumerate(LOCALITIES):
            if i > 0:
                await page.wait_for_timeout(INTER_LOCALITY_DELAY)
            print(f'Scraping {locality}...')
            data = await scrape_locality(page, locality)
            if data is not None:
                locality_bytes[locality] = data

        stop_event.set()
        watchdog.cancel()
        try:
            await watchdog
        except asyncio.CancelledError:
            pass
        await browser.close()

    if not locality_bytes:
        print('No localities downloaded — nothing to write.')
        return

    # Validate schema against the first successful download
    first_data = next(iter(locality_bytes.values()))
    actual_headers = extract_headers(first_data)
    error = validate_columns(actual_headers, EXPECTED_COLUMNS)
    if error:
        print(error)
        return

    print('Writing to Google Sheet...')
    for locality, data in locality_bytes.items():
        row_values = parse_row(data)
        new_date_val = row_values[DATE_COL]
        tab = open_sheet(SHEET_ID, locality)
        print(f'  {locality}:', end=' ')
        write_locality_data(tab, row_values, new_date_val)

    print('Done.')


if __name__ == '__main__':
    asyncio.run(main())
