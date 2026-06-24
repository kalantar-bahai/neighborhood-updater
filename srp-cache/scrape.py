"""
SRP cache scraper — authenticates with onlinesrp.org, downloads two Excel
reports, parses neighbourhood-level data, and writes it to two tabs in the
SRP cache Google Sheet.

    Devotionals tab  ← Locations → Clusters → Focus Neighbourhoods
    Education tab    ← Reports → General Reports → Institute Reports →
                        All Educational Core → By Region → By Focus Neighbourhood

Usage:
    python scrape.py

Set SRP_USERNAME, SRP_PASSWORD, and TOTP_SECRET in .env to authenticate
automatically.  If any are missing the browser opens for manual login.
"""

import asyncio
import io
from pathlib import Path

import openpyxl
from playwright.async_api import async_playwright

from srp_lib.auth import load_credentials, authenticate, totp_watchdog
from srp_lib.browser import launch_browser, download_excel
from srp_lib.browser import click as srp_click
from srp_lib.sheets import open_sheet

REGION       = 'Atlantic'
TIMEOUT  = 60_000  # ms

SHEET_ID         = '1w8eRljld_O4vkSPNwRuKWO6YL7uJIJXZ3AtqWUPi5Eo'
TAB_DEVOTIONALS  = 'Devotionals'
TAB_EDUCATION    = 'Education'
DATA_START_ROW   = 2  # row 1 will be a header written by this script

EXPORTS_DIR      = Path(__file__).parent / 'exports'


# ── Navigation ────────────────────────────────────────────────────────────────

async def set_scope(page):
    btn = page.locator('#dropdownLocationSelector')
    await btn.wait_for(state='visible')
    await btn.click()
    await page.wait_for_timeout(500)
    await page.get_by_text(REGION, exact=True).first.click()
    await page.wait_for_timeout(500)


async def scrape_devotionals(page):
    print('Scraping devotionals report...')
    await srp_click(page, page.get_by_role('link', name='Locations'))
    await srp_click(page, page.get_by_role('button', name='Clusters'))
    await srp_click(page, page.get_by_role('link', name='Focus Neighbourhoods'))
    return await download_excel(page, EXPORTS_DIR, timeout=TIMEOUT)


async def scrape_education(page):
    print('Scraping education report...')
    await srp_click(page, page.get_by_role('link', name='Reports'))
    await srp_click(page, page.get_by_role('button', name='General Reports'))
    await srp_click(page, page.get_by_role('link', name='Institute Reports'))
    await srp_click(page, page.get_by_role('button', name='All Educational Core'))
    await srp_click(page, page.get_by_role('button', name='By Region'))
    await srp_click(page, page.get_by_role('link', name='By Focus Neighbourhood'))
    return await download_excel(page, EXPORTS_DIR, timeout=TIMEOUT)


# ── Excel parsing ─────────────────────────────────────────────────────────────

def xlsx_to_rows(xlsx_bytes):
    """Return (headers, data_rows) from the first sheet of an xlsx file."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    return list(rows[0]), [list(r) for r in rows[1:]]


def parse_devotionals(xlsx_bytes):
    # Headers: Name, Latin Name, Locality, Electoral Unit, Cluster,
    #          Group of Clusters, Subregion, Region, Group of Regions,
    #          National Community, Devotional Meetings: No./Att./FoF., Comments
    headers, rows = xlsx_to_rows(xlsx_bytes)
    col = {h: i for i, h in enumerate(headers)}
    result = []
    for r in rows:
        name = r[col['Name']]
        if not name:
            continue
        result.append({
            'name':             name,
            'latin_name':       r[col['Latin Name']]       or '',
            'locality':         r[col['Locality']]         or '',
            'electoral_unit':   r[col['Electoral Unit']]   or '',
            'cluster':          r[col['Cluster']]          or '',
            'group_of_clusters':r[col['Group of Clusters']]or '',
            'subregion':        r[col['Subregion']]        or '',
            'region':           r[col['Region']]           or '',
            'group_of_regions': r[col['Group of Regions']] or '',
            'national_community':r[col['National Community']] or '',
            'dev_act':          r[col['Devotional Meetings: No.']]  or '',
            'dev_part':         r[col['Devotional Meetings: Att.']] or '',
            'dev_fof':          r[col['Devotional Meetings: FoF.']] or '',
            'comments':         r[col['Comments']]         or '',
        })
    return result


def parse_education(xlsx_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    # Row 1: merged group headers (Children's Classes | JYG | SC | All Ed)
    # Row 2: sub-column labels (No. / Participation / unlabeled-FoF per group)
    # Row 3+: data
    # Col layout:
    #  0: Focus Neighbourhood
    #  1: CC No.   2: CC Participation   3: CC FoF (unlabeled in row 2)
    #  4: JYG No.  5: JYG Participation  6: JYG FoF (unlabeled)
    #  7: SC No.   8: SC Participation   9: SC FoF (unlabeled)
    # 10: Total No. of Facilitators
    # 11-13: All Ed totals (skip — computed by web app)
    result = []
    for r in all_rows[2:]:
        name = r[0]
        if not name:
            continue
        result.append({
            'name':         name,
            'cc_act':       r[1]  or '',
            'cc_part':      r[2]  or '',
            'cc_fof':       r[3]  or '',
            'jyg_act':      r[4]  or '',
            'jyg_part':     r[5]  or '',
            'jyg_fof':      r[6]  or '',
            'sc_act':       r[7]  or '',
            'sc_part':      r[8]  or '',
            'sc_fof':       r[9]  or '',
            'facilitators': r[10] or '',
        })
    return result


# ── Google Sheet writer ───────────────────────────────────────────────────────

def write_tab(tab_name, headers, records):
    sheet = open_sheet(SHEET_ID, tab_name)
    sheet.clear()
    sheet.update(values=[headers] + records, range_name='A1')
    print(f'  Wrote {len(records)} rows to "{tab_name}" tab.')


def write_devotionals(records):
    headers = [
        'Name', 'Latin Name', 'Locality', 'Electoral Unit', 'Cluster',
        'Group of Clusters', 'Subregion', 'Region', 'Group of Regions',
        'National Community', 'Dev Active', 'Dev Participants', 'Dev FoF', 'Comments',
    ]
    rows = [[
        r.get('name',''),             r.get('latin_name',''),
        r.get('locality',''),         r.get('electoral_unit',''),
        r.get('cluster',''),          r.get('group_of_clusters',''),
        r.get('subregion',''),        r.get('region',''),
        r.get('group_of_regions',''), r.get('national_community',''),
        r.get('dev_act',''),          r.get('dev_part',''),
        r.get('dev_fof',''),          r.get('comments',''),
    ] for r in records]
    write_tab(TAB_DEVOTIONALS, headers, rows)


def write_education(records):
    headers = [
        'Name',
        'CC Active', 'CC Participants', 'CC FoF',
        'JYG Active', 'JYG Participants', 'JYG FoF',
        'SC Active', 'SC Participants', 'SC FoF',
        'Facilitators',
    ]
    rows = [[
        r.get('name',''),
        r.get('cc_act',''),  r.get('cc_part',''),  r.get('cc_fof',''),
        r.get('jyg_act',''), r.get('jyg_part',''), r.get('jyg_fof',''),
        r.get('sc_act',''),  r.get('sc_part',''),  r.get('sc_fof',''),
        r.get('facilitators',''),
    ] for r in records]
    write_tab(TAB_EDUCATION, headers, rows)


# ── Main ──────────────────────────────────────────────────────────────────────

async def main():
    creds = load_credentials()
    EXPORTS_DIR.mkdir(exist_ok=True)

    async with async_playwright() as pw:
        browser, context = await launch_browser(pw)
        page = await context.new_page()
        page.set_default_timeout(TIMEOUT)
        page.set_default_navigation_timeout(TIMEOUT)

        await page.goto(creds['url'])
        await page.wait_for_load_state('domcontentloaded')
        await page.wait_for_timeout(2_000)

        if 'login' in page.url.lower():
            if creds['username'] and creds['password'] and creds['totp_secret']:
                await authenticate(page, creds['username'], creds['password'], creds['totp_secret'])
            else:
                print('Not logged in. Set credentials in .env, or log in manually and press Enter.')
                input()

        stop_event = asyncio.Event()
        watchdog = asyncio.create_task(totp_watchdog(page, stop_event, creds['totp_secret']))

        try:
            await set_scope(page)
            dev_bytes = await scrape_devotionals(page)
            edu_bytes = await scrape_education(page)
        finally:
            stop_event.set()
            watchdog.cancel()
            try:
                await watchdog
            except asyncio.CancelledError:
                pass

        await browser.close()

    dev_records = parse_devotionals(dev_bytes)
    edu_records = parse_education(edu_bytes)

    print('Writing to Google Sheet...')
    write_devotionals(dev_records)
    write_education(edu_records)
    print('Done.')


if __name__ == '__main__':
    asyncio.run(main())
